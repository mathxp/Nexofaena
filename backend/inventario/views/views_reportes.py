# ======================
# DJANGO CORE IMPORTS
# ======================
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Sum, F, Q
from django.db.models.functions import TruncMonth

# ======================
# MODELOS
# ======================
from inventario.models import (
    Insumo,
    Movimiento,
    Consumo,
    Prediccion,
    Compra,
    DetalleCompra,
)

# ======================
# SERVICIOS (LÓGICA NEGOCIO)
# ======================
from ..services import (
    promedio_consumo_diario,
    calcular_dias_stock,
    calcular_score,
    detectar_consumo_anormal
)

# ======================
# LIBRERÍAS EXTERNAS
# ======================
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment,
    Border, Side
)
from openpyxl.drawing.image import Image

import matplotlib.pyplot as plt
from io import BytesIO

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet

import json


# =========================================================
# 📊 REPORTE INVENTARIO (HTML)
# =========================================================
@login_required
def reporte_inventario(request):
    data = [
        {
            'nombre': i.nombre,
            'stock': i.stock_actual,
            'minimo': i.stock_minimo,
            'estado': "OK" if i.stock_actual > i.stock_minimo else "BAJO STOCK"
        }
        for i in Insumo.objects.all()
    ]

    return render(request, 'reportes/inventario.html', {'data': data})


# =========================================================
# 📊 EXPORTAR INVENTARIO A EXCEL (BÁSICO)
# =========================================================
def exportar_inventario_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Inventario"

    headers = ["Insumo", "Stock Actual", "Stock Mínimo", "Estado"]
    ws.append(headers)

    for col in ws[1]:
        col.font = Font(bold=True)
        col.alignment = Alignment(horizontal="center")

    for i in Insumo.objects.all():
        estado = "OK" if i.stock_actual > i.stock_minimo else "BAJO STOCK"
        ws.append([i.nombre, i.stock_actual, i.stock_minimo, estado])

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=reporte_inventario.xlsx'

    wb.save(response)
    return response


# =========================================================
# 📄 EXPORTAR INVENTARIO A PDF
# =========================================================
def exportar_inventario_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=reporte_inventario.pdf'

    doc = SimpleDocTemplate(response)
    elements = []

    styles = getSampleStyleSheet()
    elements.append(Paragraph("Reporte de Inventario - HDA", styles['Title']))

    data = [["Insumo", "Stock", "Mínimo", "Estado"]]

    for i in Insumo.objects.all():
        estado = "OK" if i.stock_actual > i.stock_minimo else "BAJO STOCK"
        data.append([i.nombre, str(i.stock_actual), str(i.stock_minimo), estado])

    table = Table(data)

    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.darkblue),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BACKGROUND', (0,1), (-1,-1), colors.whitesmoke),
    ]))

    elements.append(table)
    doc.build(elements)

    return response


# =========================================================
# 📦 REPORTE MOVIMIENTOS (FILTROS AVANZADOS)
# =========================================================
@login_required
def reporte_movimientos(request):
    movimientos = Movimiento.objects.select_related('insumo', 'usuario', 'lote').all()

    # Filtros
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    tipo = request.GET.get('tipo')
    usuario = request.GET.get('usuario')
    vencimiento_inicio = request.GET.get('vencimiento_inicio')
    vencimiento_fin = request.GET.get('vencimiento_fin')
    estado_vencimiento = request.GET.get('estado_vencimiento')

    if fecha_inicio:
        movimientos = movimientos.filter(fecha__date__gte=fecha_inicio)

    if fecha_fin:
        movimientos = movimientos.filter(fecha__date__lte=fecha_fin)

    if tipo:
        movimientos = movimientos.filter(tipo=tipo)

    if usuario:
        movimientos = movimientos.filter(usuario__id=usuario)

    if vencimiento_inicio:
        movimientos = movimientos.filter(lote__fecha_vencimiento__gte=vencimiento_inicio)

    if vencimiento_fin:
        movimientos = movimientos.filter(lote__fecha_vencimiento__lte=vencimiento_fin)

    hoy = timezone.now().date()

    if estado_vencimiento == 'vigente':
        movimientos = movimientos.filter(lote__fecha_vencimiento__gt=hoy)

    elif estado_vencimiento == 'por_vencer':
        movimientos = movimientos.filter(
            lote__fecha_vencimiento__gt=hoy,
            lote__fecha_vencimiento__lte=hoy + timezone.timedelta(days=7)
        )

    elif estado_vencimiento == 'vencido':
        movimientos = movimientos.filter(lote__fecha_vencimiento__lt=hoy)

    usuarios = Movimiento.objects.values('usuario__id', 'usuario__username').distinct()
    insumos = Insumo.objects.all()

    filtros = {
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'tipo': tipo,
        'usuario': usuario,
        'vencimiento_inicio': vencimiento_inicio,
        'vencimiento_fin': vencimiento_fin,
        'estado_vencimiento': estado_vencimiento
    }

    return render(request, 'reportes/movimientos.html', {
        'movimientos': movimientos.order_by('-fecha'),
        'usuarios': usuarios,
        'insumos': insumos,
        'filtros': filtros
    })


# =========================================================
# 📊 ANÁLISIS DE CONSUMO
# =========================================================
@login_required
def analisis_consumo(request):
    datos = (
        Consumo.objects.values('insumo__nombre')
        .annotate(total=Sum('cantidad'))
        .order_by('-total')
    )

    return render(request, 'reportes/consumo.html', {
        'nombres': [d['insumo__nombre'] for d in datos],
        'totales': [d['total'] for d in datos]
    })


# =========================================================
# 🔮 PREDICCIÓN DE CONSUMO
# =========================================================
def generar_prediccion(request):
    predicciones = []

    for insumo in Insumo.objects.all():
        promedio = promedio_consumo_diario(insumo)
        dias = calcular_dias_stock(insumo)
        score = calcular_score(insumo)
        anormal = detectar_consumo_anormal(insumo)

        if dias <= 3:
            accion = "COMPRAR URGENTE"
        elif dias <= 7:
            accion = "PLANIFICAR COMPRA"
        else:
            accion = "OK"

        predicciones.append({
            'fecha': str(timezone.now().date()),
            'insumo': {'nombre': insumo.nombre},
            'consumo_estimado': round(promedio, 2),
            'dias_restantes': dias,
            'accion': accion,
            'score': score,
            'anormal': anormal
        })

    return render(request, 'reportes/prediccion.html', {
        'predicciones': predicciones
    })


# =========================================================
# 📊 EXPORTAR EXCEL AVANZADO + GRÁFICO
# =========================================================
@login_required
def exportar_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Inventario"

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ['Nombre', 'Stock', 'Stock Mínimo', 'Categoría', 'Área']
    sheet.append(headers)

    for col_num, col_name in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    insumos = Insumo.objects.all()

    for row_num, insumo in enumerate(insumos, start=2):
        sheet.append([
            insumo.nombre,
            insumo.stock_actual,
            insumo.stock_minimo,
            insumo.categoria.nombre,
            insumo.area.nombre
        ])

    nombres = [i.nombre for i in insumos]
    stocks = [i.stock_actual for i in insumos]

    plt.figure(figsize=(8, 4))
    plt.bar(nombres, stocks)
    plt.xticks(rotation=45, ha='right')
    plt.title("Stock de Insumos")
    plt.tight_layout()

    img_buffer = BytesIO()
    plt.savefig(img_buffer, format='png')
    img_buffer.seek(0)

    img = Image(img_buffer)
    sheet.add_image(img, "G2")

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="inventario_pro.xlsx"'

    workbook.save(response)
    return response


# =========================================================
# 📊 DASHBOARD FINANCIERO
# =========================================================
@login_required
def dashboard_financiero(request):

    total_gastado = Compra.objects.aggregate(total=Sum('total'))['total'] or 0

    gasto_mensual = (
        Compra.objects.annotate(mes=TruncMonth('fecha'))
        .values('mes')
        .annotate(total=Sum('total'))
        .order_by('mes')
    )

    gasto_proveedor = (
        Compra.objects.values('proveedor__nombre')
        .annotate(total=Sum('total'))
        .order_by('-total')
    )

    gasto_insumo = (
        DetalleCompra.objects.values('insumo__nombre')
        .annotate(total=Sum(F('cantidad') * F('precio_unitario')))
        .order_by('-total')[:10]
    )

    meses = [g['mes'].strftime("%Y-%m") if g['mes'] else "Sin fecha" for g in gasto_mensual]
    totales_mes = [float(g['total'] or 0) for g in gasto_mensual]

    proveedores = [g['proveedor__nombre'] or "Sin proveedor" for g in gasto_proveedor]
    totales_proveedor = [float(g['total'] or 0) for g in gasto_proveedor]

    insumos = [g['insumo__nombre'] or "Sin nombre" for g in gasto_insumo]
    totales_insumo = [float(g['total'] or 0) for g in gasto_insumo]

    context = {
        'total_gastado': float(total_gastado),
        'meses': json.dumps(meses),
        'totales_mes': json.dumps(totales_mes),
        'proveedores': json.dumps(proveedores),
        'totales_proveedor': json.dumps(totales_proveedor),
        'insumos': json.dumps(insumos),
        'totales_insumo': json.dumps(totales_insumo),
    }

    return render(request, 'reportes/dashboard_financiero.html', context)